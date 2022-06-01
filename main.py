import xlsxwriter

import openpyxl as op
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import math
import statistics

from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score

import statsmodels.api as sm
from scipy.stats import t
from scipy.stats import f
from scipy.stats import chi2

from scipy import stats
from scipy.stats import spearmanr

import tkinter as tk
from tkinter import *
import os

# creating a window
tsk = Tk()
tsk.title("Econometrics - seminar 2")

tsk.configure(width=400, height=210)
tsk.configure(bg='lightgray')

lab_1 = Label(tsk, text="Enter your excel file name")
lab_1.place(x=10, y=30)
a = Entry(tsk)
a.place(x=240, y=30)

lab_2 = Label(tsk, text="Enter the number of independent variables")
lab_2.place(x=10, y=60)
b = Entry(tsk)
b.place(x=240, y=60)
b.get()

lab_3 = Label(tsk, text="After saving your values to the excel file, please, push the next button")
lab_3.place(x=10, y=120)


# creating an excel file
def func1():
    wb = Workbook()
    ws = wb.active

    ws_rename = wb['Sheet']
    ws_rename.title = "Values"
    sh1 = wb["Values"]

    ws1 = wb.create_sheet("Predictions", 1)
    ws1.title = "Predictions"
    sh2 = wb["Predictions"]

    # ws1 = wb.create_sheet("Results", 2)
    # ws1.title = "Results"
    # sh3 = wb["Results"]
    #
    # ws1 = wb.create_sheet("Models comparison", 3)
    # ws1.title = "Models comparison"
    # sh4 = wb["Models comparison"]
    #
    # ws1 = wb.create_sheet("Linear trend", 4)
    # ws1.title = "Linear trend"
    # sh5 = wb["Linear trend"]

    wb.save(filename=str(a.get() + ".xlsx"))

    sh1.cell(column=1, row=1).value = "Put your data in here, rectangular form, save the file, and exit"

    sh1.cell(column=1, row=2).value = "Y"

    for i in range(1, int(b.get()) + 1):
        sh1.cell(column=1 + i, row=2).value = "X" + str(i)

    sh2.cell(column=1, row=1).value = 'Place expected values of indeces for linear trend, Y1 will adjust after ' \
                                      'clicking "Linear Trend" button '
    sh2.cell(column=4, row=1).value = 'Place expected values of arguments for regression model analysis, Y will ' \
                                      'adjust after clicking "The next button" '

    sh2.cell(column=1, row=2).value = "INDEX"
    sh2.cell(column=2, row=2).value = "Y1"

    sh2.cell(column=4, row=2).value = "Y"

    for i in range(1, int(b.get()) + 1):
        sh2.cell(column=4 + i, row=2).value = "X" + str(i)

    # sh3.cell(column=1, row=1).value = 'Results of multiple regression analysis will appear here after clicking ' \
    #                                   '"Multiple Regression Analysis"'
    #
    # sh4.cell(column=1, row=1).value = 'Models comparison will appear here after clicking "Models Comparison"'
    #
    # sh5.cell(column=1, row=1).value = 'Results on linear trend analysis will appear here after clicking "Linear Trend"'

    wb.save(filename=str(a.get() + ".xlsx"))

    os.startfile(str(a.get() + ".xlsx"))


# calculations
def func0():
    wb = op.load_workbook(str(a.get() + ".xlsx"))
    ws = wb.active

    df = pd.read_excel(str(a.get() + ".xlsx"), sheet_name='Values', header=1)
    df2 = pd.read_excel(str(a.get() + ".xlsx"), sheet_name="Predictions", header=1)
    y = df["Y"]
    x = df.drop("Y", axis=1)
    x = sm.add_constant(x)
    x_pred = df2.drop(df2.columns[[0, 1, 2, 3]], axis=1)
    x_pred = sm.add_constant(x_pred, has_constant='add')
    model = sm.OLS(y, x).fit()
    model_sk = LinearRegression().fit(x, y)
    y_pred = model.predict()
    y_av = statistics.mean(y)
    p = int(b.get())
    dof = len(y) - p - 1
    # alpha = float(c.get())
    # confidence = 1 - alpha
    confidence = 0.95

    x_new = np.zeros([len(y), int(b.get())])
    for i in range(1, int(b.get()) + 1):
        x_new[:, i - 1] = df["X" + str(i)]
    # x_np, y_np = np.array(x), np.array(y)
    return df, df2, y, x, x_pred, model, model_sk, y_pred, y_av, p, dof, confidence, x_new


def multiple_dfs(df_list, sheets, file_name, spaces):
    writer = pd.ExcelWriter(file_name, engine='openpyxl', mode="a")
    row = 0
    for dataframe in df_list:
        dataframe.to_excel(writer, sheet_name=sheets, startrow=row, startcol=0, index=False, header=False)
        row = row + len(dataframe.index) + spaces + 1
    writer.save()



def func2():
    df, df2, y, x, x_pred, model, model_sk, y_pred, y_av, p, dof, confidence, x_new = func0()

    # a) find a regression equation of Y on Х1 and X2 and explain the meaning of regression coefficients b0, b1, b2
    pr = model.params
    g = str(round((pr[0]), 4))
    for i in range(1, len(pr)):
        if pr[i] > 0:
            g += str("+" + str(round((pr[i]), 4)) + "x" + str(i))
        elif pr[i] < 0:
            g += str(str(round((pr[i]), 4)) + "x" + str(i))
        elif pr[i] == 0:
            g = g
    # print(g)

    # b) estimate the tightness and the direction of the relation between variables Х1, X2 and Y by computing multiple
    # correlation coefficient
    corr = df.corr()
    fig = plt.figure()
    ax = fig.add_subplot(111)
    cax = ax.matshow(corr, cmap='coolwarm', vmin=-1, vmax=1)
    fig.colorbar(cax)
    ticks = np.arange(0, len(df.columns), 1)
    ax.set_xticks(ticks)
    plt.xticks(rotation=90)
    ax.set_yticks(ticks)
    ax.set_xticklabels(df.columns)
    ax.set_yticklabels(df.columns)

    SSR = np.sum([(i - y.mean()) ** 2 for i in y_pred])
    SST = np.sum([(i - y.mean()) ** 2 for i in y])
    SSE_array = np.subtract(y, y_pred)
    SSE = np.sum([i ** 2 for i in SSE_array])

    r2 = model_sk.score(x, y)
    r = math.sqrt(r2)

    if abs(r) > 0.7:
        res1 = "strong connection of arguments and y"
    elif abs(r) > 0.5:
        res1 = "medium connection of arguments and y"
    else:
        res1 = "weak connection of arguments and y"
    if r > 0:
        res2 = "the variables are mutually directed"
    else:
        res2 = "the variables are oppositely directed"

    # print(r)
    # print(corr)
    # plt.show()

    # c) estimate the significance of the regression equation of Y on Х1, X2 by F-statistics at level α=0.05
    f_crit = np.abs(f.ppf(confidence, p, dof))
    F = (r2 / (1 - r2)) * (dof / p)

    if F < f_crit:
        res3 = "null hypothesis (variance x = variance y) is accepted; arguments influence is not significant"
    else:
        res3 = "null hypothesis (variance x = variance y) is rejected; arguments influence is significant"

    # d) estimate the significance of regression coefficients b0, b1, b2 by t-statistics at level α=0.05
    t_crit = np.abs(t.ppf((1 - confidence) / 2, dof))

    S2 = SSE / dof

    xt = np.matrix.transpose(x.to_numpy())
    mult1 = np.dot(xt, x)
    inv = np.linalg.inv(mult1)
    mult2 = np.dot(inv, xt)

    b_array = np.dot(mult2, y)

    res4 = []
    for i in range(0, len(b_array)):
        if b_array[i] < t_crit:
            res4.append(f"null hypothesis (β{i} = 0) is accepted; β{i} is not significant")
        else:
            res4.append(f"null hypothesis (β{i} = 0) is rejected, β{i} is significant")
    # print(b_array)

    # e) find determination coefficient and explain its meaning
    r2 = model_sk.score(x, y)
    if r2 > 0.7:
        res10 = "the model is relevant"
    else:
        res10 = "the model is hardly relevant"

    # f) estimate 95% confidence interval for the average company’s profit, given that X1 = 11, X2 = 8
    x_predt = np.matrix.transpose(x_pred.to_numpy())
    mult3 = np.dot(x_pred, inv)
    mult4 = np.dot(mult3, x_predt)
    # this is useless, i found formula in module

    y_new = model.predict(x_pred)
    conf1 = np.array([y_new - t_crit * math.sqrt(S2 * mult4), y_new + t_crit * math.sqrt(S2 * mult4)])

    # g) estimate 95% confidence interval for the individual value of company’s profit, given that X1 = 11, X2 = 8
    conf2 = np.array([y_new - t_crit * math.sqrt(S2 * (1 + mult4)), y_new + t_crit * math.sqrt(S2 * (1 + mult4))])

    # h) find an interval estimator for regression coefficients β0, β1, β2
    t_crit = np.abs(t.ppf((1 - confidence) / 2, dof))
    varb = []
    for i in range(0, int(b.get()) + 1):
        varb.append(math.sqrt(S2 * inv[i, i]))

    T = []
    for i in range(0, int(b.get()) + 1):
        T.append(abs(b_array[i]) / varb[i])

    conf3 = []
    for i in range(0, int(b.get()) + 1):
        conf3.append([b_array[i] - varb[i] * t_crit, b_array[i] + varb[i] * t_crit])

    # i) find an interval estimator for the error’s variance with 0.95 confidence
    chi2_1 = np.abs(chi2.ppf(1 - (1 - confidence) / 2, dof))
    chi2_2 = np.abs(chi2.ppf((1 - confidence) / 2, dof))
    conf4 = np.array([S2 * dof / chi2_1, S2 * dof / chi2_2])
    # print(conf4)

    # Creating Excel Writer Object from Pandas

    # writer = pd.ExcelWriter('test.xlsx')
    # workbook = writer.book
    # worksheet = workbook.add_worksheet('Validation')
    # writer.sheets['Validation'] = worksheet
    # df3.to_excel(writer, index=False, header=False, startrow=0, startcol=0)
    # df4.to_excel(writer, index=False, header=False, startrow=len(df3)+1, startcol=0)

    # 2. a) apply the Spearman rank correlation test to assess heteroscedasticity at a 5% significance level for
    counter = 0
    err = []
    while counter < int(b.get()):
        model11 = sm.OLS(y, sm.add_constant(x_new[:, counter].reshape(-1, 1))).fit()
        err.append((np.subtract(model11.predict(), y)) ** 2)
        counter += 1
    # print(err)

    x_ranks = []
    for i in range(1, int(b.get()) + 1):
        x_ranks.append(df['X' + str(i)].rank().to_numpy())
    # print(x_ranks)

    err_ranks = []
    for i in range(0, int(b.get())):
        err_df = round(pd.DataFrame(err[i]), 10)
        err_ranks.append(err_df.rank().to_numpy())
    # print(err_ranks)

    d_y = []
    for i in range(0, int(b.get())):
        d_y.append(np.subtract(x_ranks[i], np.concatenate(err_ranks[i], axis=None)) ** 2)
    # print(d_y)

    # more accurate result, but..
    # s_coef = []
    # for i in range(0, int(b.get())):
    #     s_coef.append(spearmanr(x_new[:, i], err[i]))
    # print(s_coef)

    r_xe = []
    for i in range(0, int(b.get())):
        r_xe.append((1 - 6 * sum(d_y[i]) / (len(y) * (len(y) ** 2 - 1))))
    # print(r_xe)

    t_obs_xe = []
    for i in range(0, int(b.get())):
        t_obs_xe.append(r_xe[i] * math.sqrt(len(y) - 2) / math.sqrt(1 - r_xe[i] ** 2))
    # print(t_obs_xe)

    res5 = []
    for i in range(0, len(t_obs_xe)):
        if t_obs_xe[i] < t_crit:
            res5.append("homoscedasticity")
        else:
            res5.append("heteroscedasticity")

    # b) apply the Goldfeld-Quandt test to assess heteroscedasticity at a 5% significance level for both x1 and x2
    S2_array = []
    for i in range(0, int(b.get())):
        S2_array.append(sum(err[i] ** 2) / dof)
    # print(S2_array)
    F_obs_gold = []
    for i in range(len(S2_array)):
        for j in range(len(S2_array)):
            if j > i:
                F_obs_gold.append(S2_array[j] / S2_array[i])
    # F_obs_gold = S2_array[0] / S2_array[1]
    F_crit_gold = np.abs(f.ppf(confidence, dof, dof))
    # print("F-crit =", F_crit_gold, "F-obs =", F_obs_gold)

    df3 = pd.DataFrame()
    df3.insert(0, 'a', ["a) Regression equation:", g, None, "b) Correlation coefficient:", r, res1, res2,
                        "Correlation matrix:"])
    df4 = pd.DataFrame(corr)
    df5 = pd.DataFrame()
    df5.insert(0, 'a', ["c) F-statistics:", "F critical =", "F observed =", None])
    df5.insert(1, 'b', [None, f_crit, F, res3])
    df6 = pd.DataFrame()
    df6.insert(0, 'a', ["d) T-statistics:"])
    df6.insert(1, 'b', [None])
    for i in range(0, len(b_array)):
        df6 = df6.append({'a': f"T{i} observed =", 'b': b_array[i]}, ignore_index=True)
        df6 = df6.append({'a': None, 'b': res4[i]}, ignore_index=True)
    df7 = pd.DataFrame()
    df7.insert(0, 'a', ["e) Determination coefficient:", r2, res10])
    df8 = pd.DataFrame()
    df8.insert(0, 'a', ["f) Confidence interval for the average y", conf1[0],
                        "average y of population will fall within this interval with probability of alpha and "
                        "additional conditions"])
    df8.insert(1, 'b', [None, conf1[1], None])
    df9 = pd.DataFrame()
    df9.insert(0, 'a', ["g) Confidence interval for the individual value of the number of y:", conf2[0],
                        "individual y of population will fall within this interval with probability of alpha and "
                        "additional conditions"])
    df9.insert(1, 'b', [None, conf2[1], None])
    df10 = pd.DataFrame()
    df10.insert(0, 'a', ["h) Interval estimators for regression coefficients"])
    df10.insert(1, 'b', [None])
    for i in range(0, len(conf3)):
        df10.append({'a': f"Interval estimator for regression coefficient β{i} =", 'b': None}, ignore_index=True)
        df10.append({'a': conf3[i][0], 'b': conf3[i][1]}, ignore_index=True)
    df11 = pd.DataFrame()
    df11.insert(0, 'a', ["i) Interval estimator for the error’s variance:", conf4[0],
                         "true value of the variance of the population is within this interval with 95% probability"])
    df11.insert(1, 'b', [None, conf4[1], None])

    df12 = pd.DataFrame()
    df12.insert(0, 'a', ["2. a) Spearman rank correlation test", "T critical ="])
    df12.insert(1, 'b', [None, t_crit])
    for i in range(0, len(t_obs_xe)):
        df12 = df12.append({'a': t_obs_xe[i], 'b': res5[i]}, ignore_index=True)
    df13 = pd.DataFrame()
    df13.insert(0, 'a', ["b) Goldfeld-Quandt test", "F-crit =", F_crit_gold, "F-obs =", F_obs_gold])
    # seems like y variance and x[i] variance should be compared (while i compared x[i] and x[j] values)

    # df_list = [df3, df4, df5, df6, df7, df8, df9, df10, df11, df12, df13]
    list_please = [pd.concat([df3, df4, df5, df6, df7, df8, df9, df10, df11, df12, df13], sort=False)]
    multiple_dfs(list_please, sheets='Results', file_name=str(a.get()+".xlsx"), spaces=0)


def func3():
    df, df2, y, x, x_pred, model, model_sk, y_pred, y_av, p, dof, confidence, x_new = func0()
    # 3. a) According to the table, construct an empirical regression equation for: a) the power function у=β0*x_1
    # ^(β_1 )*x_2^(β_2 )*ε
    x_ln = np.log(x_new)
    x_ln = sm.add_constant(x_ln)
    y_ln = np.log(y.to_numpy())
    model_power = sm.OLS(y_ln, x_ln).fit()

    pr_power = model_power.params
    pr_power[0] = math.exp(pr_power[0])
    g1 = str(round((pr_power[0]), 4))
    for i in range(1, len(pr_power)):
        if pr_power[i] > 0:
            g1 += str("*" + "x" + str(i) + "^" + str(round((pr_power[i]), 4)))
        elif pr_power[i] < 0:
            g1 += str("*" + "x" + str(i) + "^(" + str(round((pr_power[i]), 4)) + ")")
        elif pr_power[i] == 0:
            g1 = g1
    # print(g1)
    # print(pr_power)
    err_pr = []
    yyy = np.zeros(len(y))
    for i in range(0, len(y)):
        yyy[i] = pr_power[0]
        for j in range(0, int(b.get())):
            yyy[i] = yyy[i]*x_new[i, j]**pr_power[j+1]
    for i in range(0, len(y)):
        err_pr.append(np.subtract(yyy[i], y[i]))
    sum_pr = np.zeros(len(y))
    for i in range(0, len(y)):
        sum_pr[i] = abs(err_pr[i]/y[i])
    A1 = sum(sum_pr)/len(sum_pr)
    r21 = LinearRegression().fit(x_ln, y_ln).score(x_ln, y_ln)

    # print("power model:", g1)

    # b) the equilateral hyperbola у = β0 + β1 / x1 + β2 / x2 + ε
    x_hyp = np.zeros([len(y), int(b.get())])

    for i in range(1, int(b.get()) + 1):
        x_hyp[:, i - 1] = df["X" + str(i)]
    for i in range(0, len(y)):
        x_hyp[i] = 1 / x_hyp[i]
    x_hyp = sm.add_constant(x_hyp)
    model_hyperbola = sm.OLS(y, x_hyp).fit()

    pr_hyp = model_hyperbola.params
    g2 = str(round((pr_hyp[0]), 4))
    for i in range(1, len(pr_hyp)):
        if pr_hyp[i] > 0:
            g2 += str("+", str(round((pr_hyp[i]), 4)) + "/" + "x" + str(i))
        elif pr_hyp[i] < 0:
            g2 += str(str(round((pr_hyp[i]), 4)) + "/" + "x" + str(i))
        elif pr_hyp[i] == 0:
            g2 = g2

    err_hyp = []
    for i in range(0, len(y)):
        err_hyp.append(np.subtract(model_hyperbola.predict()[i], y[i]))
    sum_hyp = np.zeros(len(y))
    for i in range(0, len(y)):
        sum_hyp[i] = abs(err_hyp[i] / y[i])
    A2 = sum(sum_hyp) / len(sum_hyp)
    r22 = LinearRegression().fit(x_hyp, y).score(x_hyp, y)

    # print("hyperbolic:", g2)

    # c) the exponential function у = β0 * е ^ (β_1 x_1+β_2 x_2) * ε;
    model_exp = sm.OLS(y_ln, x).fit()
    pr_exp1 = model_exp.params
    pr_exp1[0] = math.exp(pr_exp1[0])
    g3 = str(round((pr_exp1[0]), 4)) + "e^("
    for i in range(1, len(pr_exp1)):
        if pr_exp1[i] > 0:
            g3 += str("+" + str(round((pr_exp1[i]), 4)) + "x" + str(i))
        elif pr_exp1[i] < 0:
            g3 += str(str(round((pr_exp1[i]), 4)) + "x" + str(i))
        elif pr_exp1[i] == 0:
            g3 = g3
    g3 = g3 + ")"

    err_exp1 = []
    yyy_exp1 = np.zeros(len(y))

    power = np.zeros(len(y))
    for i in range(0, len(y)):
        for j in range(0, int(b.get())):
            power[i] += x_new[i, j] * pr_exp1[j + 1]

    for i in range(0, len(y)):
        yyy_exp1[i] = pr_exp1[0]*math.exp(power[i])

    for i in range(0, len(y)):
        err_exp1.append(np.subtract(yyy_exp1[i], y[i]))

    sum_exp1 = np.zeros(len(y))
    for i in range(0, len(y)):
        sum_exp1[i] = abs(err_exp1[i] / y[i])
    A3 = sum(sum_exp1) / len(sum_exp1)
    r23 = LinearRegression().fit(x, y_ln).score(x, y_ln)
    # print("exponential model 1:", g3)

    # d) the semi - logarithmic function у = β 0 + β1lnx1 + β2lnx2 + ε;
    model_log = sm.OLS(y, x_ln).fit()
    pr_log = model_log.params
    g4 = str(round((pr_log[0]), 4))
    for i in range(1, len(pr_log)):
        if pr_log[i] > 0:
            g4 += str("+" + str(round((pr_log[i]), 4)) + "lnx" + str(i))
        elif pr_log[i] < 0:
            g4 += str(str(round((pr_log[i]), 4)) + "lnx" + str(i))
        elif pr_log[i] == 0:
            g4 = g4

    err_log = []
    for i in range(0, len(y)):
        err_log.append(np.subtract(model_log.predict()[i], y[i]))
    sum_log = np.zeros(len(y))
    for i in range(0, len(y)):
        sum_log[i] = abs(err_log[i] / y[i])
    A4 = sum(sum_log) / len(sum_log)
    r24 = LinearRegression().fit(x_ln, y).score(x_ln, y)
    # print("semi-logarithmic model:", g4)

    # e) the inverse function у = 1 / (β0 + β1x1 + β2x2 + ε);
    y_hyp = []
    for i in range(0, len(y)):
        y_hyp.append(y[i])
    for i in range(0, len(y)):
        y_hyp[i] = 1 / y_hyp[i]
    model_inv = sm.OLS(y_hyp, x).fit()

    pr_inv = model_inv.params
    g5 = "1/(" + str(round((pr_inv[0]), 4))
    for i in range(1, len(pr_inv)):
        if pr_inv[i] > 0:
            g5 += str("+" + str(round((pr_inv[i]), 4)) + "x" + str(i))
        elif pr_inv[i] < 0:
            g5 += str(str(round((pr_inv[i]), 4)) + "x" + str(i))
        elif pr_inv[i] == 0:
            g5 = g5
    g5 = g5 + ")"

    err_inv = []
    yyy_inv = np.zeros(len(y))

    denom = np.zeros(len(y))
    for i in range(0, len(y)):
        denom[i] = pr_inv[0]
        for j in range(0, int(b.get())):
            denom[i] += x_new[i, j] * pr_inv[j + 1]

    for i in range(0, len(y)):
        yyy_inv[i] = 1/denom[i]

    for i in range(0, len(y)):
        err_inv.append(np.subtract(yyy_inv[i], y[i]))

    sum_inv = np.zeros(len(y))
    for i in range(0, len(y)):
        sum_inv[i] = abs(err_inv[i] / y[i])
    A5 = sum(sum_inv) / len(sum_inv)
    r25 = LinearRegression().fit(x, y_hyp).score(x, y_hyp)

    # print("inverse model:", g5)

    # f) the function у = β0 + β1√(x_1) + β2√(x_2) + ε;
    x_sqrt = np.zeros([len(y), int(b.get())])
    for i in range(1, int(b.get()) + 1):
        x_sqrt[:, i - 1] = df["X" + str(i)]
    for i in range(0, len(y)):
        for j in range(0, int(b.get())):
            x_sqrt[i, j] = math.sqrt(x_sqrt[i, j])
    x_sqrt = sm.add_constant(x_sqrt)
    model_sqrt = sm.OLS(y, x_sqrt).fit()

    pr_sqrt = model_sqrt.params
    # print(pr_sqrt)
    g6 = str(round((pr_sqrt[0]), 4))
    for i in range(1, len(pr_sqrt)):
        if pr_sqrt[i] > 0:
            g6 += str("+" + str(round((pr_sqrt[i]), 4)) + "sqrt(x" + str(i) + ")")
        elif pr_sqrt[i] < 0:
            g6 += str(str(round((pr_sqrt[i]), 4)) + "sqrt(x" + str(i) + ")")
        elif pr_sqrt[i] == 0:
            g6 = g6
    g6 = g6 + ")"

    err_sqrt = []
    for i in range(0, len(y)):
        err_sqrt.append(np.subtract(model_sqrt.predict()[i], y[i]))
    sum_sqrt = np.zeros(len(y))
    for i in range(0, len(y)):
        sum_sqrt[i] = abs(err_sqrt[i] / y[i])
    A6 = sum(sum_sqrt) / len(sum_sqrt)
    r26 = LinearRegression().fit(x_sqrt, y).score(x_sqrt, y)

    # print("sqrt model:", g6)

    # g) the exponential function у = β0β_1 ^ (x_1) β_2 ^ (x_2) * ε;
    pr_exp2 = sm.OLS(y_ln, x).fit().params
    for i in range(0, len(pr_exp2)):
        pr_exp2[i] = math.exp(pr_exp2[i])
    g7 = str(round((pr_exp2[0]), 4))
    for i in range(1, len(pr_exp2)):
        if pr_exp2[i] > 0:
            g7 += str("*" + str(round((pr_exp2[i]), 4)) + "^x" + str(i))
        elif pr_exp2[i] < 0:
            g7 += str("*" + "(" + str(round((pr_exp2[i]), 4)) + ")^x" + str(i))
        elif pr_exp2[i] == 0:
            g7 = g7

    err_exp2 = []
    yyy_exp2 = np.zeros(len(y))
    for i in range(0, len(y)):
        yyy_exp2[i] = pr_exp2[0]
        for j in range(0, int(b.get())):
            yyy_exp2[i] = yyy_exp2[i] * pr_exp2[j + 1]**x_new[i, j]
    for i in range(0, len(y)):
        err_exp2.append(np.subtract(yyy_exp2[i], y[i]))
    sum_exp2 = np.zeros(len(y))
    for i in range(0, len(y)):
        sum_exp2[i] = abs(err_exp2[i] / y[i])
    A7 = sum(sum_exp2) / len(sum_exp2)
    r27 = LinearRegression().fit(x, y_ln).score(x, y_ln)

    # print("exponential model 2:", g7)

    # h) estimate for each of the models the determination coefficient, the average error approximation, and choose
    # the best model.
    err_ = []
    for i in range(0, len(y)):
        err_.append(np.subtract(model.predict()[i], y[i]))
    sum_ = np.zeros(len(y))
    for i in range(0, len(y)):
        sum_[i] = abs(err_[i] / y[i])
    A0 = sum(sum_) / len(sum_)
    r20 = model_sk.score(x, y)

    A_array = [A1, A2, A3, A4, A5, A6, A7, A0]
    max_A = max(A_array)
    r2_array = [r21, r22, r23, r24, r25, r26, r27, r20]
    max_r2 = max(r2_array)

    df3 = pd.DataFrame()
    df3.insert(0, 'a', ["a) the power function у=β0*x_1^(β_1 )*x_2^(β_2 )*ε]", g1, "Average error approximation =",
               "R2 =", None, "b) the equilateral hyperbola у = β0 + β1 / x1 + β2 / x2 + ε", g2,
                        "Average error approximation =", "R2 =", None,
                        "c) the exponential function у = β0 * е ^ (β_1 x_1+β_2 x_2) * ε", g3,
                        "Average error approximation =", "R2 =", None,
                        "d) the semi - logarithmic function у = β 0 + β1lnx1 + β2lnx2 + ε", g4,
                        "Average error approximation =", "R2 =", None,
                        "e) the inverse function у = 1 / (β0 + β1x1 + β2x2 + ε)", g5,
                        "Average error approximation =", "R2 =", None,
                        "f) the function у = β0 + β1√(x_1) + β2√(x_2) + ε", g6,
                        "Average error approximation =", "R2 =", None,
                        "g) the exponential function у = β0β_1 ^ (x_1) β_2 ^ (x_2) * ε", g7,
                        "Average error approximation =", "R2 =", None,
                        "h) initial model", "Average error approximation =", "R2 ="])
    df3.insert(1, 'b', [None, None, A1, r21, None, None, None, A2, r22, None, None, None, A3, r23, None, None, None, A4,
                        r24, None, None, None, A5, r25, None, None, None, A6, r26, None, None, None, A7, r27, None,
                        None, A0, r20])

    df_list = [df3]
    multiple_dfs(df_list, sheets='Models comparison', file_name=str(a.get()+".xlsx"), spaces=0)


def func4():
    df, df2, y, x, x_pred, model, model_sk, y_pred, y_av, p, dof, confidence, x_new = func0()
    # 4. 3. According to the table for the time series уt:
    # a) find the equation of the non-random component (trend), assuming the trend is linear;
    index_array = []
    for i in range(0, len(y)):
        index_array.append(i + 1)

    model_lt = sm.OLS(y.to_numpy(), sm.add_constant(index_array)).fit()
    pr = model_lt.params
    g = str(round((pr[0]), 4))
    for i in range(1, len(pr)):
        if pr[i] > 0:
            g += str("+" + str(round((pr[i]), 4)) + "x" + str(i))
        elif pr[i] < 0:
            g += str(str(round((pr[i]), 4)) + "x" + str(i))
        elif pr[i] == 0:
            g = g
    rsq = model_lt.rsquared
    if abs(rsq) > 0.7:
        res1 = "strong connection"
    elif abs(rsq) > 0.5:
        res1 = "medium connection"
    else:
        res1 = "weak connection"
    if rsq > 0:
        res2 = "mutual direction"
    else:
        res2 = "opposite direction"
    # print(model_lt.params, model_lt.rsquared)

    # b) identify, at the significance level 0.05, the presence of autocorrelation of errors using the Durbin-Watson
    # criterion;
    errors = np.zeros([len(y)])
    for i in range(0, len(y)):
        errors[i] = y[i] - model_lt.predict()[i]

    errors2 = np.zeros([len(y)])
    for i in range(0, len(y)):
        errors2[i] = errors[i] ** 2

    errorst = np.zeros([len(y)])
    errorst[0] = 0
    for i in range(1, len(y)):
        errorst[i] = errors[i - 1]

    errors_num = np.zeros([len(y)])
    for i in range(0, len(y)):
        errors_num[i] = (errors[i] - errorst[i]) ** 2
    errors_num[0] = 0

    durb_wats = sum(errors_num) / sum(errors2)
    # print(durb_wats)

    # print(sm.stats.stattools.durbin_watson(model_lt.resid, axis=0))

    # c) find, with confidence 0.95, an interval estimate of the true variance of errors, assuming the trend is linear;
    chi2_11 = np.abs(chi2.ppf(1 - (1 - confidence) / 2, len(y) - 2))
    chi2_22 = np.abs(chi2.ppf((1 - confidence) / 2, len(y) - 2))
    RSS_array = np.subtract(y, model_lt.predict())
    RSS = np.sum([i ** 2 for i in RSS_array])
    ss2 = RSS / (len(y) - 2)
    conf_linear = np.array([ss2 * (len(y) - 2) / chi2_11, ss2 * (len(y) - 2) / chi2_22])

    # print(conf_linear)

    # d) find the autocorrelation coefficient (for the lag τ=1,2,3);
    lag_2 = np.zeros([len(y)])
    lag_2[0] = 0
    for i in range(1, len(y)):
        lag_2[i] = errorst[i - 1]

    lag_3 = np.zeros([len(y)])
    lag_3[0] = 0
    for i in range(1, len(y)):
        lag_3[i] = lag_2[i - 1]
    # print(errors)
    # print(errorst)
    # print(lag_2)
    # print(lag_3)
    # print(np.corrcoef(errors, errorst)[0, 1], np.corrcoef(errors, lag_2)[1, 0], np.corrcoef(errors, lag_3)[1, 0])
    errors_br = [(i - errors.mean()) for i in errors]
    errors_br2 = np.sum([(i - errors.mean()) ** 2 for i in errors])
    errorst_br = [(i - errorst.mean()) for i in errorst]
    errorst_br2 = np.sum([(i - errorst.mean()) ** 2 for i in errorst])
    lag_2_br = [(i - lag_2.mean()) for i in lag_2]
    lag_2_br2 = np.sum([(i - lag_2.mean()) ** 2 for i in lag_2])
    lag_3_br = [(i - lag_3.mean()) for i in lag_3]
    lag_3_br2 = np.sum([(i - lag_3.mean()) ** 2 for i in lag_3])
    correl_1 = np.sum(np.multiply(errors_br, errorst_br)) / math.sqrt(errors_br2 * errorst_br2)
    correl_2 = np.sum(np.multiply(errors_br, lag_2_br)) / math.sqrt(errors_br2 * lag_2_br2)
    correl_3 = np.sum(np.multiply(errors_br, lag_3_br)) / math.sqrt(errors_br2 * lag_3_br2)
    # print(correl_2, correl_3, "ya hochoo plakat")

    # e) find, with confidence 0.95, an interval estimation of the regression coefficient β1, assuming the trend is
    # linear;
    index_arr = sm.add_constant(index_array)
    index_arrayt = np.matrix.transpose(index_arr)
    mult1_y = np.dot(index_arrayt, index_arr)
    inv_y = np.linalg.inv(mult1_y)

    varb0_y = math.sqrt(ss2 * inv_y[0, 0])
    varb1_y = math.sqrt(ss2 * inv_y[1, 1])

    t_crit_y = np.abs(t.ppf((1 - confidence) / 2, len(y) - 2))

    confb0_y = np.array([model_lt.params[0] - t_crit_y * varb0_y, model_lt.params[0] + t_crit_y * varb0_y])
    confb1_y = np.array([model_lt.params[1] - t_crit_y * varb1_y, model_lt.params[1] + t_crit_y * varb1_y])

    # print(confb0_y, confb1_y)

    # f) evaluate with confidence 0.95 the significance of the pair regression coefficient using the t-test, assuming
    # the trend is linear;

    T0_y = abs(model_lt.params[0]) / varb0_y
    T1_y = abs(model_lt.params[1]) / varb1_y

    if T0_y < t_crit_y:
        res3 = "null hypothesis (β0 = 0) is accepted; β0 is not significant"
    else:
        res3 = "null hypothesis (β0 = 0) is accepted; β0 is significant"

    if T1_y < t_crit_y:
        res4 = "null hypothesis (β1 = 0) is accepted; β1 is not significant"
    else:
        res4 = "null hypothesis (β1 = 0) is accepted; β1 is significant"

    # print("t crit:", t_crit_y, "T0:", T0_y, "T1:", T1_y)

    # g) find a point estimate and with the confidence of 0.95 interval estimation of the forecast of the average
    # (individual) value of the company’s profit at the time t = 11 (eleventh year), assuming the trend is linear;
    index_array_n = sm.add_constant(df2['INDEX'].to_numpy(), has_constant='add')
    y_lin_n = model_lt.predict(index_array_n)

    mult2_y = np.dot(index_array_n, inv_y)
    mult3_y = np.dot(mult2_y, np.matrix.transpose(index_array_n))

    conf1_y = np.array([y_lin_n[0] - t_crit_y * math.sqrt(ss2 * mult3_y),
                        y_lin_n[0] + t_crit_y * math.sqrt(ss2 * mult3_y)])
    conf2_y = np.array([y_lin_n[0] - t_crit_y * math.sqrt(ss2 * (1 + mult3_y)),
                        y_lin_n[0] + t_crit_y * math.sqrt(ss2 * (1 + mult3_y))])
    # print(conf1_y, conf2_y)

    # h) check with the confidence of 0.95 the significance of pair regression using the F-test, assuming the trend
    # is linear.
    y_pred_lin = model_lt.predict()
    ESS_y = np.sum([(i - y.to_numpy().mean()) ** 2 for i in y_pred_lin])

    f_crit_y = np.abs(f.ppf(confidence, 1, len(y) - 2))
    F_y = (ESS_y / 1) / (RSS / (len(y) - 2))

    if F_y < f_crit_y:
        res5 = "null hypothesis (variance x = variance y) is accepted; influence is not significant"
    else:
        res5 = "null hypothesis (variance x = variance y) is rejected; influence is significant"

    # print("f crit:", f_crit_y, "F:", F_y)
    df3 = pd.DataFrame()
    df3.insert(0, 'a', ["a) Regression equation:", g, "Correlation coefficient:", rsq, res1, res2, None,
                        "b) Durbin-Watson criterion", durb_wats, None])
    df4 = pd.DataFrame()
    df4.insert(0, 'a', ["c) Interval estimator for the error’s variance:", conf_linear[0],
                        "true value of the variance of the population is within this interval with 95% probability"])
    df4.insert(1, 'b', [None, conf_linear[1], None])
    df5 = pd.DataFrame()
    df5.insert(0, 'a', ["d) find the autocorrelation coefficient (for the lag τ=1,2,3)", correl_1, correl_2, correl_3])
    df6 = pd.DataFrame()
    df6.insert(0, 'a', ["e) Interval estimator for regression coefficient β0:", confb0_y[0],
                        "Interval estimator for regression coefficient β1:", confb1_y[0]])
    df6.insert(1, 'b', [None, confb0_y[1], None, confb1_y[1]])
    df7 = pd.DataFrame()
    df7.insert(0, 'a', ["f) Significance of regression coefficients", "T critical:", t_crit_y, "T0:", T0_y, res3, "T1:",
                        T1_y, res4])
    df8 = pd.DataFrame()
    df8.insert(0, 'a', ["g) Confidence interval for the average y", conf1_y[0],
                        "average y of population will fall within this interval with probability of alpha and "
                        "additional conditions"])
    df8.insert(1, 'b', [None, conf1_y[1], None])
    df9 = pd.DataFrame()
    df9.insert(0, 'a', ["Confidence interval for the individual value of the number of y:", conf2_y[0],
                        "individual y of population will fall within this interval with probability of alpha and "
                        "additional conditions"])
    df9.insert(1, 'b', [None, conf2_y[1], None])
    df10 = pd.DataFrame()
    df10.insert(0, 'a', ["h) Significance of pair regression", "F critical:", f_crit_y, "F observed:", F_y, res5])

    list_please = [pd.concat([df3, df4, df5, df6, df7, df8, df9, df10], sort=False)]
    multiple_dfs(list_please, sheets='Linear trend', file_name=str(a.get() + ".xlsx"), spaces=0)


def func5():
    os.startfile(str(a.get() + ".xlsx"))


# buttons
new_excel = Button(tsk, text="Create an excel file", command=func1)
new_excel.place(x=10, y=90)

open_excelfile = Button(tsk, text="Open the file", command=func5)
open_excelfile.place(x=160, y=90)

calc1_button = Button(tsk, text="Multiple Regression Analysis", command=func2)
calc1_button.place(x=10, y=150)

calc2_button = Button(tsk, text="Models Comparison", command=func3)
calc2_button.place(x=265, y=150)

calc3_button = Button(tsk, text="Linear Trend", command=func4)
calc3_button.place(x=180, y=150)

mainloop()
